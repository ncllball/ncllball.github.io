#!/usr/bin/env python3
"""
build_packets.py - NCLL 2026 Medical Waiver Packet Builder

Two-phase workflow:
  Phase 1 - classify:
    python build_packets.py --classify
    Reads rosters + form files, classifies every form, writes MANIFEST.xlsx.
    Open MANIFEST.xlsx in Excel, review flagged forms, and set
    the Override column to "Acceptable" for any you want to force-accept.

  Phase 2 - build:
    python build_packets.py --build
    Reads MANIFEST.xlsx (respecting your overrides), generates one PDF packet
    per team and a SUMMARY.xlsx.

Requirements:
    pip install openpyxl pillow numpy opencv-python reportlab pypdf
"""

import argparse
import io
import os
import re
import subprocess
import sys
from collections import defaultdict
from pathlib import Path

# Force UTF-8 output on Windows so special characters in names print cleanly
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ── dependency imports ────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("Missing: pip install openpyxl")

try:
    import numpy as np
    import cv2
    from PIL import Image, ImageOps
except ImportError:
    sys.exit("Missing: pip install pillow numpy opencv-python")

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        Image as RLImage, PageBreak, HRFlowable,
    )
    from reportlab.lib.enums import TA_LEFT, TA_CENTER
except ImportError:
    sys.exit("Missing: pip install reportlab")

try:
    import pypdf
except ImportError:
    sys.exit("Missing: pip install pypdf")

# ── paths — adjust if your folders move ──────────────────────────────────────
BASE_DIR       = Path(r"C:\Tools\ncllball.github.io\.github\projects\waiver-builder")
WAIVERS_DIR    = BASE_DIR / "blank_forms"
TEAMS_DIR      = BASE_DIR / "rosters"
OUTPUT_DIR     = BASE_DIR / "coach_packets"
MANIFEST_PATH  = OUTPUT_DIR / "MANIFEST.xlsx"
SUMMARY_PATH   = OUTPUT_DIR / "SUMMARY.xlsx"
EMPTY_WAIVER   = WAIVERS_DIR / "medical.release.waiver.2026.empty.pdf"

# ── classification thresholds ─────────────────────────────────────────────────
BLURRY_THRESHOLD    = 50.0    # Laplacian variance below this = blurry
BLACK_BOX_THRESHOLD = 15.0    # Mean pixel brightness (0-255) below this = black box

# ── status constants ──────────────────────────────────────────────────────────
ACCEPTABLE    = "Acceptable"
BLACK_BOX_ST  = "Black Box"
BLURRY_ST     = "Blurry"
WRONG_FORM_ST = "Wrong Form"
INCOMPLETE_ST = "Incomplete"
NOT_SUBMITTED = "Not Submitted"

REASON_LABELS = {
    BLACK_BOX_ST:  "Black box — form is unreadable",
    BLURRY_ST:     "Blurry — key fields cannot be read",
    WRONG_FORM_ST: "Wrong form — not a Little League Medical Release Form",
    INCOMPLETE_ST: "Incomplete — one or more required fields are missing",
    NOT_SUBMITTED: "Not submitted — no form on file",
}

VALID_DIV_PROGRAMS = {
    "TB":  ["BB"],
    "KCP": ["BB"],
    "A":   ["BB", "SB"],
    "AA":  ["BB", "SB"],
    "AAA": ["BB", "SB"],
    "MAJ": ["BB", "SB"],
    "JUN": ["BB", "SB"],
    "SEN": ["BB", "SB"],
}

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".webp"}

PAGE_W, PAGE_H = letter  # 612 × 792 points


# ── utilities ─────────────────────────────────────────────────────────────────

def sanitize(s: str) -> str:
    s = re.sub(r'[<>:"/\\|?*]', "", s)
    return re.sub(r"\s+", " ", s).strip()


def normalize(s: str) -> str:
    """Lowercase, split camelCase, remove punctuation, collapse whitespace."""
    # Split camelCase/PascalCase before lowercasing (e.g. WandlerBaxter -> Wandler Baxter)
    s = re.sub(r"([a-z])([A-Z])", r"\1 \2", s)
    # Strip punctuation used as separators (/, -, _, space, apostrophe, period)
    s = re.sub(r"[\s\-_/'`.]+", " ", s.lower())
    return s.strip()


def name_in_stem(player_name: str, stem: str) -> bool:
    """True if player's full name appears in the filename stem."""
    parts = player_name.split()
    if len(parts) < 2:
        return normalize(player_name) in normalize(stem)
    first, last = parts[0], parts[-1]
    fn = normalize(stem)
    return (
        f"{normalize(first)} {normalize(last)}" in fn
        or normalize(player_name) in fn
    )


# ── roster loading ────────────────────────────────────────────────────────────

def _parse_division_name(s: str) -> str:
    """Map a 'Division Name:' string to our short code."""
    sl = s.lower()
    # NOTE: "tball" is a substring of "softball" — use word-boundary check instead
    if re.search(r"\bt[\s\-]?ball\b", sl) or "tee-ball" in sl or "tee ball" in sl:
        return "TB"
    if "kindergarten" in sl or "coach pitch" in sl or "kcp" in sl:
        return "KCP"
    if "senior" in sl:
        return "SEN"
    if "junior" in sl:
        return "JUN"
    if "major" in sl:
        return "MAJ"
    if "aaa" in sl or " aaa" in sl:
        return "AAA"
    if "aa" in sl:
        return "AA"
    if re.search(r"\ba\b", sl) or sl.startswith("a -") or sl.startswith("a division"):
        return "A"
    return s.split()[0].upper()


def load_rosters() -> list[dict]:
    """
    Parse the master roster file (2026.final.ncll.roster.xlsx).
    Columns: Program Name(0), Division Name(1), AcctFirst(2), AcctLast(3),
             PlayerFirst(4), PlayerLast(5), Gender(6), DOB(7), ..., TeamName(17)
    Returns a list of team dicts:
      {team_name, division, program, coach_name, coach_email, players: [str]}
    """
    roster_path = TEAMS_DIR / "2026.final.ncll.roster.xlsx"
    if not roster_path.exists():
        # Fallback: old per-division waivers_*.xlsx files
        return _load_rosters_legacy()

    try:
        wb = openpyxl.load_workbook(roster_path, data_only=True)
    except Exception as e:
        print(f"  [WARN] Cannot open {roster_path.name}: {e}")
        return _load_rosters_legacy()

    ws = wb.active
    teams_dict: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        prog_raw  = str(row[0]).strip() if row[0] else ""
        div_raw   = str(row[1]).strip() if row[1] else ""
        p_first   = str(row[4]).strip() if row[4] else ""
        p_last    = str(row[5]).strip() if row[5] else ""
        team_name = str(row[17]).strip() if row[17] else ""

        if not (p_first and p_last and team_name) or team_name == "Unallocated":
            continue

        program  = "SB" if "SOFTBALL" in prog_raw.upper() else "BB"
        division = _parse_division_name(div_raw)

        if team_name not in teams_dict:
            teams_dict[team_name] = {
                "team_name":   team_name,
                "division":    division,
                "program":     program,
                "coach_name":  "Coaching Staff",
                "coach_email": "",
                "players":     [],
            }
        teams_dict[team_name]["players"].append(f"{p_first} {p_last}")

    teams = list(teams_dict.values())

    # Filter to valid divisions
    valid = []
    for t in teams:
        if t["division"] in VALID_DIV_PROGRAMS and t["program"] in VALID_DIV_PROGRAMS[t["division"]]:
            valid.append(t)
        else:
            print(f"  [SKIP] {t['team_name']} — division '{t['division']}' not in scope")
    return valid


def _load_rosters_legacy() -> list[dict]:
    """Fallback: parse old-format waivers_*.xlsx files."""
    teams = []
    for xlsx_path in sorted(TEAMS_DIR.glob("waivers_*.xlsx")):
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        except Exception as e:
            print(f"  [WARN] Cannot open {xlsx_path.name}: {e}")
            continue
        ws = wb.active
        current_program = ""
        current_division = ""
        current_team: dict | None = None
        mode = None
        for row in ws.iter_rows(values_only=True):
            v0 = str(row[0]).strip() if row[0] else ""
            if v0.startswith("Program Name:"):
                prog_str = v0.split(":", 1)[1].upper()
                current_program = "SB" if "SOFTBALL" in prog_str else "BB"
            elif v0.startswith("Division Name:"):
                current_division = _parse_division_name(v0.split(":", 1)[1].strip())
            elif v0.startswith("Team Name:"):
                if current_team:
                    teams.append(current_team)
                current_team = {"team_name": v0.split(":", 1)[1].strip(),
                                "division": current_division, "program": current_program,
                                "coach_name": "Coaching Staff", "coach_email": "", "players": []}
                mode = None
            elif v0 == "Team Players":
                mode = "players"
            elif v0 == "Team Personnel":
                mode = "personnel"
            elif mode == "players" and current_team:
                first = str(row[1]).strip() if row[1] else ""
                last  = str(row[2]).strip() if row[2] else ""
                if first and last and first != "Player First Name":
                    current_team["players"].append(f"{first} {last}")
            elif mode == "personnel" and current_team:
                role  = str(row[1]).strip() if row[1] else ""
                first = str(row[2]).strip() if row[2] else ""
                last  = str(row[3]).strip() if row[3] else ""
                email = str(row[4]).strip() if row[4] else ""
                if role == "Head Coach" and first and last:
                    current_team["coach_name"]  = f"{first} {last}"
                    current_team["coach_email"] = email
        if current_team:
            teams.append(current_team)
    valid = []
    for t in teams:
        if t["division"] in VALID_DIV_PROGRAMS and t["program"] in VALID_DIV_PROGRAMS[t["division"]]:
            valid.append(t)
    return valid

    return valid


# ── form file discovery ───────────────────────────────────────────────────────

def team_folder(team_name: str) -> Path | None:
    """Find the subfolder in WAIVERS_DIR that matches this team name."""
    exact = WAIVERS_DIR / team_name
    if exact.is_dir():
        return exact

    # Fuzzy: find best-scoring folder by normalized name overlap
    norm_target = normalize(team_name)
    best, best_score = None, 0
    for d in WAIVERS_DIR.iterdir():
        if not d.is_dir():
            continue
        norm_d = normalize(d.name)
        # Score by longest common token sequence
        target_tokens = set(norm_target.split())
        d_tokens      = set(norm_d.split())
        shared = len(target_tokens & d_tokens)
        if shared > best_score:
            best_score = shared
            best = d

    return best if best_score >= 2 else None


def find_player_files(player_name: str, folder: Path) -> list[Path]:
    """Return all image/PDF files anywhere in WAIVERS_DIR whose name contains
    the player's name.

    The Sports Connect team assignments don't always match the rosters,
    so we search the entire blank_forms tree rather than a single team folder.
    """
    results = []
    for sub in WAIVERS_DIR.iterdir():
        if not sub.is_dir():
            continue
        for f in sub.iterdir():
            if not f.is_file():
                continue
            ext = f.suffix.lower()
            if ext not in IMAGE_EXTS and ext not in (".pdf", ".docx", ".doc"):
                continue
            if name_in_stem(player_name, f.stem):
                results.append(f)
    return sorted(results)


# ── image classification ──────────────────────────────────────────────────────

def correct_orientation(img: Image.Image) -> Image.Image:
    """Apply EXIF rotation; then rotate to portrait if landscape."""
    try:
        img = ImageOps.exif_transpose(img)
    except Exception:
        pass
    if img.width > img.height:
        img = img.rotate(-90, expand=True)
    return img


def _shrink_pil(img: Image.Image, max_px: int = 1500) -> Image.Image:
    """Downscale a PIL image for analysis only — does not affect the PDF copy."""
    w, h = img.size
    longest = max(w, h)
    if longest <= max_px:
        return img
    scale = max_px / longest
    return img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)


def _cap_for_pdf(img: Image.Image, max_px: int = 1500) -> Image.Image:
    """Cap resolution for PDF embedding — 1500px is plenty for 8.5x11 at 150dpi."""
    w, h = img.size
    longest = max(w, h)
    if longest <= max_px:
        return img
    scale = max_px / longest
    return img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)


def _to_cv2(img: Image.Image) -> "np.ndarray":
    return cv2.cvtColor(np.array(img.convert("RGB")), cv2.COLOR_RGB2BGR)


def _shrink_for_analysis(cv_img: "np.ndarray", max_px: int = 1500) -> "np.ndarray":
    """Downscale so longest side <= max_px before analysis (saves memory)."""
    h, w = cv_img.shape[:2]
    longest = max(h, w)
    if longest <= max_px:
        return cv_img
    scale = max_px / longest
    return cv2.resize(cv_img, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_AREA)


def _is_black_box(cv_img: "np.ndarray") -> bool:
    gray = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY)
    return float(gray.mean()) < BLACK_BOX_THRESHOLD


def _is_blurry(cv_img: "np.ndarray") -> bool:
    gray = cv2.cvtColor(cv_img, cv2.COLOR_BGR2GRAY)
    lap = cv2.Laplacian(gray, cv2.CV_64F)
    return float(lap.var()) < BLURRY_THRESHOLD


def classify_file(path: Path) -> tuple[str, str, list[Image.Image]]:
    """
    Returns (status, detail, corrected_images).
    corrected_images is populated only for ACCEPTABLE forms.
    """
    ext = path.suffix.lower()

    # ── Word document — accept as-is, cannot analyze visually ────────────────
    if ext in (".docx", ".doc"):
        return ACCEPTABLE, "Word document (unanalyzed)", []

    # ── PDF ──────────────────────────────────────────────────────────────────
    if ext == ".pdf":
        try:
            reader = pypdf.PdfReader(str(path))
        except Exception as e:
            return BLACK_BOX_ST, f"unreadable PDF: {e}", []
        if not reader.pages:
            return BLACK_BOX_ST, "empty PDF", []
        # Try to rasterize for analysis (requires pdf2image / poppler)
        try:
            from pdf2image import convert_from_path
            images = [correct_orientation(img) for img in convert_from_path(str(path), dpi=150)]
            cv_imgs = [_to_cv2(img) for img in images]
            if all(_is_black_box(c) for c in cv_imgs):
                return BLACK_BOX_ST, "", []
            if any(_is_blurry(c) for c in cv_imgs):
                return BLURRY_ST, "", []
            return ACCEPTABLE, "", images
        except ImportError:
            # Can't analyze — include as-is and trust the user
            return ACCEPTABLE, "PDF included unanalyzed (install pdf2image for analysis)", []
        except Exception as e:
            return BLACK_BOX_ST, f"PDF rasterize failed: {e}", []

    # ── Image ─────────────────────────────────────────────────────────────────
    try:
        img = Image.open(path)
        img = correct_orientation(img)
    except Exception as e:
        return BLACK_BOX_ST, f"cannot open: {e}", []

    # Shrink a copy for analysis to avoid OOM on large images (800px is plenty for blur/black detection)
    small = _shrink_pil(img, max_px=800)
    cv_img = _to_cv2(small)

    if _is_black_box(cv_img):
        return BLACK_BOX_ST, "", []
    if _is_blurry(cv_img):
        return BLURRY_ST, "", []

    # Cap resolution for PDF embedding
    return ACCEPTABLE, "", [_cap_for_pdf(img, max_px=2000)]


def best_classification(files: list[Path]) -> tuple[str, str, list[Image.Image], Path | None]:
    """
    Evaluate all files for one player; return the best result.
    Priority: ACCEPTABLE > any non-NOT_SUBMITTED > NOT_SUBMITTED
    Returns (status, detail, images, winning_file_path).
    """
    best_status, best_detail, best_images, best_file = NOT_SUBMITTED, "", [], None

    for f in files:
        status, detail, images = classify_file(f)
        if status == ACCEPTABLE:
            return ACCEPTABLE, detail, images, f
        if best_status == NOT_SUBMITTED:
            best_status, best_detail, best_images, best_file = status, detail, images, f

    return best_status, best_detail, best_images, best_file


# ── manifest ──────────────────────────────────────────────────────────────────

MANIFEST_COLS = [
    "Team Name", "Division", "Program", "Coach Name",
    "Player Name", "File", "Auto Status", "Detail", "Override",
]

MANIFEST_INSTRUCTIONS = (
    "Review the Auto Status column for each player. "
    "If you want to force-accept a flagged form, type  Acceptable  in the Override column. "
    "Leave Override blank to use the auto classification. "
    "Save this file, then run:  python build_packets.py --build"
)


def write_manifest(rows: list[dict]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Manifest"

    # Instructions row
    ws.append([MANIFEST_INSTRUCTIONS])
    ws.merge_cells(f"A1:{chr(64 + len(MANIFEST_COLS))}1")
    ws["A1"].font       = Font(italic=True, color="555555")
    ws["A1"].alignment  = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 32

    # Header
    ws.append(MANIFEST_COLS)
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    for cell in ws[2]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data
    for r in rows:
        ws.append([
            r["team_name"], r["division"], r["program"], r["coach_name"],
            r["player_name"], r.get("file", ""), r["auto_status"],
            r.get("detail", ""), "",   # Override column — blank for user
        ])

    # Highlight non-acceptable rows
    flag_fill  = PatternFill("solid", fgColor="FFF3CD")  # amber
    ok_fill    = PatternFill("solid", fgColor="D4EDDA")  # green
    for row in ws.iter_rows(min_row=3):
        status_cell = row[6]   # "Auto Status" column (0-indexed: col G = index 6)
        fill = ok_fill if status_cell.value == ACCEPTABLE else flag_fill
        for cell in row:
            cell.fill = fill

    # Column widths
    widths = [38, 8, 8, 22, 24, 60, 14, 30, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Override column header highlight
    override_col = MANIFEST_COLS.index("Override") + 1
    ws.cell(row=2, column=override_col).fill = PatternFill("solid", fgColor="C8E6FA")

    wb.save(MANIFEST_PATH)
    print(f"\n  Manifest: {MANIFEST_PATH}")


def read_manifest_overrides() -> dict[tuple, str]:
    """
    Returns {(team_name, player_name): override_status} for rows where
    Override column is non-empty.
    """
    if not MANIFEST_PATH.exists():
        return {}
    try:
        wb   = openpyxl.load_workbook(MANIFEST_PATH, data_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        print(f"  [WARN] Cannot read manifest: {e}")
        return {}

    overrides = {}
    for row in rows[2:]:   # skip instructions + header
        if not row or not row[0]:
            continue
        team_name   = str(row[0]).strip()
        player_name = str(row[4]).strip() if row[4] else ""
        override    = str(row[8]).strip() if row[8] else ""
        if override:
            overrides[(team_name, player_name)] = override
    return overrides


# ── PDF builder ───────────────────────────────────────────────────────────────

def _style(name, **kwargs) -> ParagraphStyle:
    base = getSampleStyleSheet()["Normal"]
    return ParagraphStyle(name, parent=base, **kwargs)


H1   = _style("h1", fontSize=16, leading=20, fontName="Helvetica-Bold", spaceAfter=4)
H2   = _style("h2", fontSize=13, leading=17, fontName="Helvetica-Bold", spaceAfter=4)
H3   = _style("h3", fontSize=11, leading=14, fontName="Helvetica-Bold", spaceAfter=3)
BODY = _style("body", fontSize=10, leading=14, fontName="Helvetica", spaceAfter=6)
STEP = _style("step", fontSize=10, leading=14, fontName="Helvetica", leftIndent=20, spaceAfter=4)
META = _style("meta", fontSize=9,  leading=12, fontName="Helvetica", textColor=colors.HexColor("#666666"))


def _build_front_matter(team: dict, player_results: list[dict]) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=inch, rightMargin=inch,
        topMargin=inch, bottomMargin=inch,
    )
    elems = []

    no_form = [r for r in player_results if not r.get("images")]

    coach_name = team.get("coach_name") or "Coaching Staff"
    salutation = coach_name.split()[0] if coach_name != "Coaching Staff" else "Coaching Staff"

    # ── Cover page ────────────────────────────────────────────────────────────
    elems.append(Paragraph(f"{team['team_name']} — Medical Release Packet", H1))
    elems.append(Paragraph(f"{team['division']} | {team['program']} | 2026 Season", META))
    elems.append(Spacer(1, 0.15 * inch))
    elems.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1A3A5C")))
    elems.append(Spacer(1, 0.2 * inch))

    elems.append(Paragraph(f"Hi {salutation},", BODY))
    elems.append(Spacer(1, 0.1 * inch))

    elems.append(Paragraph(
        "Attached are the medical release forms currently on file for your players. "
        "Please follow these steps to keep them accessible during every game and practice:", BODY,
    ))
    elems.append(Spacer(1, 0.1 * inch))
    for i, step in enumerate([
        "Place all printed forms inside a <b>Ziplock bag</b>, then seal that bag inside a <b>second Ziplock bag</b>.",
        "Drop the sealed bag into your <b>ball bucket, underneath the balls</b>.",
        "<b>Point it out to your entire coaching staff</b> so everyone knows where it is in case of an emergency.",
    ], 1):
        elems.append(Paragraph(f"{i}. {step}", STEP))

    elems.append(Spacer(1, 0.15 * inch))

    # ── Disclaimer ────────────────────────────────────────────────────────────
    DISCLAIMER = _style("disclaimer", fontSize=10, leading=14, fontName="Helvetica",
                        backColor=colors.HexColor("#FFF8E1"),
                        borderPadding=(6, 8, 6, 8), spaceAfter=4)
    DISC_STEP  = _style("disc_step", fontSize=10, leading=14, fontName="Helvetica",
                        backColor=colors.HexColor("#FFF8E1"),
                        leftIndent=20, borderPadding=(0, 8, 2, 8), spaceAfter=2)

    elems.append(Paragraph(
        "<b>Please note:</b> This year's medical waivers have a few issues that coaches need "
        "to address with their parents. Go through your waivers and if you feel comfortable "
        "with having to use that waiver in an emergency situation, then that parent is done. "
        "If not, please have them fill out a new one. A blank waiver is included on page 3 "
        "of this packet for those cases.", DISCLAIMER,
    ))
    elems.append(Spacer(1, 0.08 * inch))
    elems.append(Paragraph("<b>Potential issues to look for:</b>", DISCLAIMER))
    for bullet in [
        "<b>Black pages</b> most likely mean some type of interaction with e-sign in Acrobat — probably a legitimate upload bug on the parent's end.",
        "<b>Wrong document</b> (driver's license, etc.) means the parent uploaded the wrong file during registration.",
        "<b>No waiver or multiple empty waivers</b> means the parent either did not upload one or uploaded a blank form.",
    ]:
        elems.append(Paragraph(f"\u2022 {bullet}", DISC_STEP))

    elems.append(Spacer(1, 0.15 * inch))

    # ── Closing reference to page 2 ───────────────────────────────────────────
    closing = (
        "<b>Page 2</b> of this packet contains your full team roster alongside the list of "
        "forms included. Use it to verify every player is accounted for."
    )
    if no_form:
        closing += (
            " Players listed under <i>No Form on File</i> must submit a completed waiver "
            "at the first practice — use the blank form on page 3."
        )
    elems.append(Paragraph(closing, BODY))

    doc.build(elems)
    return buf.getvalue()


def _mini_table(data: list[list], col_widths: list, font_size: int = 9) -> Table:
    """Build a compact inner table for use inside a two-column layout."""
    tbl = Table(data, colWidths=col_widths)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0), colors.HexColor("#1A3A5C")),
        ("TEXTCOLOR",      (0, 0), (-1, 0), colors.white),
        ("FONTNAME",       (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",       (0, 0), (-1, -1), font_size),
        ("FONTNAME",       (0, 1), (-1, -1), "Helvetica"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
        ("GRID",           (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",     (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 3),
    ]))
    return tbl


def _build_summary_page(team: dict, player_results: list[dict]) -> bytes:
    """
    Page 2: team roster (left column) + Forms in Packet / No Form on File (right column).
    Everything on a single page — font scales down if needed to fit.
    """
    has_form = sorted(
        [r for r in player_results if r.get("images")],
        key=lambda x: x["player_name"].split()[-1].lower(),
    )
    no_form = sorted(
        [r for r in player_results if not r.get("images")],
        key=lambda x: x["player_name"].split()[-1].lower(),
    )
    players = sorted(team.get("players", []), key=lambda n: n.split()[-1].lower())

    # ── choose font size so both columns fit on one page ──────────────────────
    # Largest list drives the row count; each row ~(font_size + 6) pts tall + header
    max_rows = max(len(players), len(has_form) + len(no_form) + (2 if no_form else 0))
    # usable height after margins + header block (~1.6 inch)
    usable_h = PAGE_H - 2 * inch - 1.6 * inch
    # estimate: header row 20pt + data rows (font+6)
    for fs in (9, 8, 7):
        row_h = fs + 6
        est_h = 20 + max_rows * row_h
        if est_h <= usable_h:
            break

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=inch, rightMargin=inch,
        topMargin=inch, bottomMargin=inch,
    )
    elems = []

    H2sm = _style("h2sm", fontSize=11, leading=14, fontName="Helvetica-Bold", spaceAfter=3)
    METAsm = _style("metasm", fontSize=8, leading=10, fontName="Helvetica",
                    textColor=colors.HexColor("#666666"))
    BODYsm = _style("bodysm", fontSize=fs, leading=fs + 3, fontName="Helvetica", spaceAfter=4)

    elems.append(Paragraph(f"{team['team_name']} — Roster &amp; Forms Summary", H2sm))
    elems.append(Paragraph(f"{team['division']} | {team['program']} | 2026 Season", METAsm))
    elems.append(Spacer(1, 0.08 * inch))
    elems.append(HRFlowable(width="100%", thickness=0.75, color=colors.HexColor("#1A3A5C")))
    elems.append(Spacer(1, 0.1 * inch))
    elems.append(Paragraph(
        "Use this page to verify every player on your roster has a waiver in this packet. "
        "Report any discrepancies to the registrar.", BODYsm,
    ))
    elems.append(Spacer(1, 0.1 * inch))

    # Column widths: left (roster) 3.2", gap 0.1", right (forms) 3.7"
    L, R = 3.2 * inch, 3.7 * inch

    # ── Left: roster ──────────────────────────────────────────────────────────
    roster_data = [["#", "Team Roster"]]
    for i, name in enumerate(players, 1):
        roster_data.append([str(i), name])
    left_tbl = _mini_table(roster_data, [0.35 * inch, L - 0.35 * inch], font_size=fs)

    # ── Right: forms in packet + no form on file ──────────────────────────────
    right_inner = io.BytesIO()   # we'll build right column as a list of flowables in a KeepTogether

    forms_data = [["Forms in This Packet"]]
    for r in has_form:
        forms_data.append([r["player_name"]])
    right_forms = _mini_table(forms_data, [R], font_size=fs)

    if no_form:
        noform_data = [["No Form on File — Collect at First Practice"]]
        for r in no_form:
            noform_data.append([r["player_name"]])
        right_noform = _mini_table(noform_data, [R], font_size=fs)
    else:
        right_noform = None

    # Pack right column into a nested table so it stays together
    right_cells = [[right_forms]]
    if right_noform:
        right_cells.append([Spacer(1, 0.1 * inch)])
        right_cells.append([right_noform])
    right_tbl = Table(right_cells, colWidths=[R])
    right_tbl.setStyle(TableStyle([("TOPPADDING", (0, 0), (-1, -1), 0),
                                   ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                                   ("LEFTPADDING", (0, 0), (-1, -1), 0),
                                   ("RIGHTPADDING", (0, 0), (-1, -1), 0)]))

    # ── Two-column outer table ─────────────────────────────────────────────────
    outer = Table(
        [[left_tbl, right_tbl]],
        colWidths=[L, R],
        hAlign="LEFT",
    )
    outer.setStyle(TableStyle([
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",  (0, 0), (0, -1),  6),   # small gap between columns
        ("RIGHTPADDING",  (1, 0), (1, -1),  0),
    ]))
    elems.append(outer)

    doc.build(elems)
    return buf.getvalue()


def _images_to_pdf_bytes(images: list[Image.Image]) -> bytes:
    """Embed one player's (orientation-corrected) images into a PDF, one image per page."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=0.25 * inch, rightMargin=0.25 * inch,
        topMargin=0.25 * inch, bottomMargin=0.25 * inch,
    )
    elems = []
    usable_w = PAGE_W - 0.5 * inch
    usable_h = PAGE_H - 0.75 * inch   # conservative — ReportLab frame is slightly smaller than margins suggest

    for i, img in enumerate(images):
        if i > 0:
            elems.append(PageBreak())

        scale  = min(usable_w / img.width, usable_h / img.height)
        draw_w = img.width  * scale
        draw_h = img.height * scale

        # JPEG avoids ReportLab's ascii85 encoding of raw pixel data — much lower memory
        img_buf = io.BytesIO()
        img.convert("RGB").save(img_buf, format="JPEG", quality=88)
        img_buf.seek(0)
        elems.append(RLImage(img_buf, width=draw_w, height=draw_h))

    doc.build(elems)
    return buf.getvalue()


def _pdf_pages_to_bytes_with_header(pdf_path: Path, player_name: str = None) -> bytes:
    """Return the PDF pages directly (no separate name header page)."""
    return pdf_path.read_bytes()


def _merge_pdfs(pdf_bytes_list: list[bytes]) -> bytes:
    writer = pypdf.PdfWriter()
    for b in pdf_bytes_list:
        reader = pypdf.PdfReader(io.BytesIO(b))
        for page in reader.pages:
            writer.add_page(page)
    buf = io.BytesIO()
    writer.write(buf)
    return buf.getvalue()


def build_team_pdf(team: dict, player_results: list[dict]) -> bytes:
    # Page 1: cover letter
    parts = [_build_front_matter(team, player_results)]

    # Page 2: roster + forms summary (one page, two columns)
    parts.append(_build_summary_page(team, player_results))

    # Page 3: blank waiver for re-collects
    if EMPTY_WAIVER.exists():
        parts.append(EMPTY_WAIVER.read_bytes())

    # Pages 4+: individual waivers sorted by last name
    has_form = sorted(
        [r for r in player_results if r.get("images")],
        key=lambda r: r["player_name"].split()[-1].lower(),
    )
    for r in has_form:
        parts.append(_images_to_pdf_bytes(r["images"]))

    merged = _merge_pdfs(parts)

    # ── Embed PDF metadata (visible in Windows Explorer > Details tab) ────────
    reader = pypdf.PdfReader(io.BytesIO(merged))
    writer = pypdf.PdfWriter()
    writer.append_pages_from_reader(reader)

    coach_name = team.get("coach_name") or ""
    player_names = ", ".join(
        sorted((r["player_name"] for r in player_results), key=lambda n: n.split()[-1].lower())
    )
    writer.add_metadata({
        "/Title":    f"{team['team_name']} — Medical Release Packet",
        "/Author":   "NCLL Registrar",
        "/Subject":  f"NCLL 2026 Medical Waivers — {team['division']} {team['program']}",
        "/Keywords": f"NCLL; 2026; medical waiver; {team['division']}; {team['program']}; {team['team_name']}; {coach_name}",
        "/Creator":  "NCLL build_packets.py",
    })

    buf = io.BytesIO()
    writer.write(buf)
    return buf.getvalue()


# ── summary spreadsheet ───────────────────────────────────────────────────────

def write_summary(all_team_results: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    headers = [
        "Team", "Division", "Program", "Coach",
        "Total Players", "Acceptable",
        "Black Box", "Blurry", "Wrong Form", "Incomplete", "Not Submitted",
    ]
    ws.append(headers)
    hf = PatternFill("solid", fgColor="1A3A5C")
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = hf
        cell.alignment = Alignment(horizontal="center")

    totals: dict[str, int] = defaultdict(int)
    for t in sorted(all_team_results, key=lambda x: (x["division"], x["program"], x["team_name"])):
        c = t["counts"]
        total = sum(c.values())
        ws.append([
            t["team_name"], t["division"], t["program"], t["coach_name"],
            total,
            c.get(ACCEPTABLE, 0),
            c.get(BLACK_BOX_ST, 0),
            c.get(BLURRY_ST, 0),
            c.get(WRONG_FORM_ST, 0),
            c.get(INCOMPLETE_ST, 0),
            c.get(NOT_SUBMITTED, 0),
        ])
        for k, v in c.items():
            totals[k] += v
        totals["__total__"] += total

    # Totals row
    ws.append([
        "TOTAL", "", "", "",
        totals["__total__"],
        totals.get(ACCEPTABLE, 0),
        totals.get(BLACK_BOX_ST, 0),
        totals.get(BLURRY_ST, 0),
        totals.get(WRONG_FORM_ST, 0),
        totals.get(INCOMPLETE_ST, 0),
        totals.get(NOT_SUBMITTED, 0),
    ])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    col_widths = [40, 8, 8, 24, 12, 12, 10, 8, 12, 12, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(SUMMARY_PATH)
    print(f"  Summary: {SUMMARY_PATH}")


# ── classify phase ────────────────────────────────────────────────────────────

def run_classify() -> list[dict]:
    """Classify all forms and write MANIFEST.xlsx. Returns manifest rows."""
    print("\nLoading rosters …")
    teams = load_rosters()
    print(f"  {len(teams)} teams in scope")

    manifest_rows = []
    grand: dict[str, int] = defaultdict(int)

    for team in teams:
        folder = team_folder(team["team_name"])
        div_prog = f"{team['division']} {team['program']}"
        print(f"\n-- {div_prog:10} {team['team_name']}")
        if not folder:
            print(f"  [WARN] No matching folder in blank_forms/")

        for player_name in team["players"]:
            files = find_player_files(player_name, folder) if folder else []
            status, detail, images, winning_file = (
                best_classification(files) if files else (NOT_SUBMITTED, "", [], None)
            )
            grand[status] += 1

            icon = "OK" if status == ACCEPTABLE else "!!"
            print(f"  {icon} {player_name}: {status}" + (f" ({detail})" if detail else ""))

            manifest_rows.append({
                "team_name":   team["team_name"],
                "division":    team["division"],
                "program":     team["program"],
                "coach_name":  team["coach_name"],
                "player_name": player_name,
                "file":        str(winning_file) if winning_file else "",
                "auto_status": status,
                "detail":      detail,
            })

    write_manifest(manifest_rows)

    print("\n" + "=" * 60)
    print("CLASSIFICATION SUMMARY")
    print("=" * 60)
    total = sum(grand.values())
    print(f"  Total players:     {total}")
    print(f"  Acceptable:        {grand.get(ACCEPTABLE, 0)}")
    print(f"  Flagged:")
    for s in [BLACK_BOX_ST, BLURRY_ST, WRONG_FORM_ST, INCOMPLETE_ST, NOT_SUBMITTED]:
        n = grand.get(s, 0)
        if n:
            print(f"    {s:<18} {n}")
    print(f"\n  Open MANIFEST.xlsx, review flagged forms, set Override = 'Acceptable'")
    print(f"  where appropriate, then run:  python build_packets.py --build")

    return manifest_rows


# ── build phase ───────────────────────────────────────────────────────────────

def run_build() -> None:
    """Read MANIFEST overrides, load forms, generate PDFs and SUMMARY.xlsx."""
    print("\nLoading rosters …")
    teams = load_rosters()

    overrides = read_manifest_overrides()
    if overrides:
        print(f"  {len(overrides)} manual override(s) found in MANIFEST.xlsx")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    all_team_results = []
    grand: dict[str, int] = defaultdict(int)

    for team in teams:
        folder   = team_folder(team["team_name"])
        div_prog = f"{team['division']} {team['program']}"
        print(f"\n── {div_prog:10} {team['team_name']}")

        player_results = []
        counts: dict[str, int] = defaultdict(int)

        for player_name in team["players"]:
            files = find_player_files(player_name, folder) if folder else []
            status, detail, images, winning_file = (
                best_classification(files) if files else (NOT_SUBMITTED, "", [], None)
            )

            # Apply manual override
            override_key    = (team["team_name"], player_name)
            override_status = overrides.get(override_key, "").strip()
            if override_status:
                # Re-load images for overridden forms if we don't have them yet
                if not images and winning_file:
                    _, _, images, _ = best_classification([winning_file])
                if not images and files:
                    _, _, images, _ = best_classification(files)
                final_status = override_status
                print(f"  ** {player_name}: {status} => OVERRIDE => {final_status}")
            else:
                final_status = status
                icon = "OK" if final_status == ACCEPTABLE else "!!"
                print(f"  {icon} {player_name}: {final_status}" + (f" ({detail})" if detail else ""))

            counts[final_status] += 1
            grand[final_status]  += 1
            player_results.append({
                "player_name":  player_name,
                "auto_status":  status,
                "final_status": final_status,
                "detail":       detail,
                "images":       images,
            })

        # Write team PDF
        out_dir = OUTPUT_DIR / f"{team['division']}-{team['program']}"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{sanitize(team['team_name'])}.pdf"

        try:
            pdf_bytes = build_team_pdf(team, player_results)
            out_path.write_bytes(pdf_bytes)
            print(f"  PDF: {out_path.relative_to(OUTPUT_DIR.parent)}")
        except Exception as e:
            import traceback
            print(f"  [ERROR] PDF failed: {e}")
            traceback.print_exc()

        all_team_results.append({
            "team_name":  team["team_name"],
            "division":   team["division"],
            "program":    team["program"],
            "coach_name": team["coach_name"],
            "counts":     dict(counts),
        })

    write_summary(all_team_results)

    print("\n" + "=" * 60)
    print("BUILD COMPLETE")
    print("=" * 60)
    total = sum(grand.values())
    resub = total - grand.get(ACCEPTABLE, 0)
    print(f"  Teams processed:     {len(all_team_results)}")
    print(f"  Total players:       {total}")
    print(f"  Acceptable forms:    {grand.get(ACCEPTABLE, 0)}")
    print(f"  Resubmissions needed: {resub}")
    for s in [BLACK_BOX_ST, BLURRY_ST, WRONG_FORM_ST, INCOMPLETE_ST, NOT_SUBMITTED]:
        n = grand.get(s, 0)
        if n:
            print(f"    {s:<18} {n}")
    print(f"\n  Output: {OUTPUT_DIR}")


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="NCLL 2026 Medical Waiver Packet Builder",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Workflow:
  1. python build_packets.py --classify
       Scans all forms, writes MANIFEST.xlsx. Open it, review flagged entries,
       type 'Acceptable' in the Override column for any you want to force-accept.

  2. python build_packets.py --build
       Reads MANIFEST.xlsx overrides, generates one PDF per team + SUMMARY.xlsx.
        """,
    )
    parser.add_argument("--classify", action="store_true", help="Phase 1: classify forms and write MANIFEST.xlsx")
    parser.add_argument("--build",    action="store_true", help="Phase 2: build PDFs using MANIFEST overrides")
    args = parser.parse_args()

    print("=" * 60)
    print("NCLL 2026 Medical Waiver Packet Builder")
    print("=" * 60)
    print(f"  Base    : {BASE_DIR}")
    print(f"  Waivers : {WAIVERS_DIR}")
    print(f"  Rosters : {TEAMS_DIR}")
    print(f"  Output  : {OUTPUT_DIR}")

    if not WAIVERS_DIR.is_dir():
        sys.exit(f"\n[ERROR] blank_forms folder not found:\n  {WAIVERS_DIR}")
    if not (TEAMS_DIR / "2026.final.ncll.roster.xlsx").exists() and not any(TEAMS_DIR.glob("waivers_*.xlsx")):
        sys.exit(f"\n[ERROR] No roster files found in:\n  {TEAMS_DIR}")

    if args.classify and args.build:
        sys.exit("[ERROR] Use --classify or --build, not both.")
    elif args.classify:
        run_classify()
    elif args.build:
        if not MANIFEST_PATH.exists():
            print(f"[WARN] No MANIFEST.xlsx found — building without overrides.")
        run_build()
    else:
        parser.print_help()
        print("\nRun --classify first to generate MANIFEST.xlsx for review.")


if __name__ == "__main__":
    main()
