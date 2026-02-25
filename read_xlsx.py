import subprocess, sys

# Try to import openpyxl; install if missing
try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

import os

folder = r'C:\Tools\ncllball.github.io\.github\projects\feb26'
files = [f for f in os.listdir(folder) if f.endswith('.xlsx')]

for fname in sorted(files):
    path = os.path.join(folder, fname)
    print(f"\n{'='*60}")
    print(f"FILE: {fname}")
    print(f"{'='*60}")
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"\n  Sheet: {sheet}  ({ws.max_row} rows x {ws.max_column} cols)")
        # Print first 60 rows
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 60:
                print(f"  ... ({ws.max_row - 60} more rows)")
                break
            # Skip completely empty rows
            if all(v is None for v in row):
                continue
            vals = [str(v) if v is not None else '' for v in row]
            print('  | ' + ' | '.join(vals))
